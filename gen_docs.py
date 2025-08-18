import os
import shutil
from openpyxl import load_workbook
from openpyxl.styles.colors import COLOR_INDEX

import yaml


DOCS_DIR = "docs"
VEHICLES_DIR = "docs/vehicles"
TEMP_DIR = "_temp"


def delete_vehicle_subdirs():
    vehicles_path = os.path.join(os.getcwd(), VEHICLES_DIR)
    if os.path.exists(vehicles_path):
        for item in os.listdir(vehicles_path):
            item_path = os.path.join(vehicles_path, item)
            if os.path.isdir(item_path):
                shutil.rmtree(item_path)


def xlsx_to_html(xlsx_file, output_file):
    print(f"Converting {xlsx_file} to {output_file}...")
    wb = load_workbook(xlsx_file, data_only=True)
    ws = wb['Connections']

    num_cols = 8
    html_rows = []

    # Create table header
    header_cells = ws[1]
    ths = []
    for i in range(num_cols):
        cell = header_cells[i] if i < len(header_cells) else None
        value = cell.value if cell.value is not None else ""
        ths.append(f"<th>{value}</th>")
    html_rows.append("<thead><tr>" + "".join(ths) + "</tr></thead>")

    # Create table body
    html_rows.append("<tbody>")
    curr_enclosure = None
    curr_connector = None
    for row in ws.iter_rows(min_row=2, values_only=False):
        tds = []
        for i in range(num_cols):
            cell = row[i] if i < len(row) else None

            value = cell.value if cell is not None and cell.value is not None else ""
            table_cell = f"<td>{value}</td>"

            if i == 0:
                if cell is not None and cell.value is not None:
                    curr_enclosure = cell.value
                if curr_enclosure:
                    value = f"{curr_enclosure}"
                else:
                    value = ""
                table_cell = f"<td><strong>{value}</strong></td>"
            elif i == 1:
                if cell is not None and cell.value is not None:
                    curr_connector = cell.value
                if curr_connector:
                    value = f"{curr_connector}"
                else:
                    value = ""
                table_cell = f"<td>{value}</td>"
            elif i == 3:
                from openpyxl.styles.colors import COLOR_INDEX
                color = None
                if cell.fill and cell.fill.start_color:
                    if cell.fill.start_color.type == "rgb":
                        color = cell.fill.start_color.rgb[2:]
                    elif cell.fill.start_color.type == "theme":
                        theme_color = cell.fill.start_color.theme
                        if theme_color in COLOR_INDEX:
                            color = COLOR_INDEX[theme_color][2:]
                if color:
                    table_cell = f'<td style="background-color: #{color};">{value}</td>'
                else:
                    table_cell = f"<td>{value}</td>"
            tds.append(table_cell)

        # Pad row if shorter than header
        while len(tds) < num_cols:
            tds.append("<td></td>")

        html_rows.append("<tr>" + "".join(tds) + "</tr>")
        html_rows.append("</tbody>")

    # Wrap table
    html_table = "<table id='excelTable' class='display' style='width:100%'>\n" + "\n".join(html_rows) + "\n</table>"

    # Save HTML
    with open(output_file, "w", encoding="utf-8") as f:
        f.write(html_table)

    print(f"Saved HTML table to {output_file}")


def copy_projects():
    projects_path = os.path.join(os.getcwd(), TEMP_DIR)
    vehicles_path = os.path.join(os.getcwd(), VEHICLES_DIR)

    print(f"Copying projects from {projects_path} to {vehicles_path}...")

    if os.path.exists(projects_path):
        for vehicle in os.listdir(projects_path):
            vehicle_src = os.path.join(projects_path, vehicle)
            print(f"Processing vehicle: {vehicle}")
            if os.path.isdir(vehicle_src):
                vehicle_dst = os.path.join(vehicles_path, vehicle)
                os.makedirs(vehicle_dst, exist_ok=True)

                for project in os.listdir(vehicle_src):
                    project_src = os.path.join(vehicle_src, project)
                    if os.path.isdir(project_src):
                        project_dst = os.path.join(vehicle_dst, project)
                        os.makedirs(project_dst, exist_ok=True)

                        pdf_name = f"{vehicle}_{project}_guide.pdf"
                        html_name = f"{vehicle}_{project}_model.html"
                        pinout_name = f"{vehicle}_{project}_connections.html"

                        for file in os.listdir(project_src):
                            src_file = os.path.join(project_src, file)
                            if file.lower().endswith(".html"):
                                dst_file = os.path.join(project_dst, html_name)
                                shutil.copy2(src_file, dst_file)
                            elif file.lower().endswith(".pdf"):
                                dst_file = os.path.join(project_dst, pdf_name)
                                shutil.copy2(src_file, dst_file)
                            elif file.lower().endswith(".xlsx") and not file.lower().startswith("~"):
                                dst_file = os.path.join(project_dst, pinout_name)
                                xlsx_to_html(src_file, dst_file)


def generate_docs_and_nav():
    nav = []

    # Walk through vehicles
    for vehicle in sorted(os.listdir(VEHICLES_DIR)):
        vehicle_path = os.path.join(VEHICLES_DIR, vehicle)
        if os.path.isdir(vehicle_path):
            vehicle_nav = []

            # Walk through projects
            for project in sorted(os.listdir(vehicle_path)):
                project_path = os.path.join(VEHICLES_DIR, vehicle, project)
                if os.path.isdir(project_path):
                    print(f"Processing project: {project} for vehicle: {vehicle}")
                    pdf_orig = next((f for f in os.listdir(project_path) if f.lower().endswith(".pdf")), None)
                    html_orig = next((f for f in os.listdir(project_path) if f.lower().endswith(".html")), None)

                    # Define new file names with underscores
                    pdf_name = f"{vehicle}_{project}_guide.pdf"
                    html_name = f"{vehicle}_{project}_model.html"
                    pinout_name = f"{vehicle}_{project}_connections.html"

                    pdf_exists = os.path.exists(os.path.join(project_path, pdf_name))
                    html_exists = os.path.exists(os.path.join(project_path, html_name))
                    pinout_exists = os.path.exists(os.path.join(project_path, pinout_name))

                    # Generate index.md for the project
                    md_content = f"# {project} Guides\n\n"

                    # Material attr_list buttons
                    if pdf_exists:
                        md_content += f"[Open PDF]({pdf_name}){{: .md-button .md-raised target=\"_blank\" }}\n\n"
                    if html_exists:
                        md_content += f"[Open 3D Model]({html_name}){{: .md-button .md-raised target=\"_blank\" }}\n\n"
                    if pinout_exists:
                        md_content += f'''
## Harness Pinouts

<div id="pinouts-table">
    {{% include "Vehicles/{vehicle}/{project}/{pinout_name}" %}}
</div>

<script src="https://code.jquery.com/jquery-3.7.1.min.js"></script>
<script src="https://cdn.datatables.net/1.13.6/js/jquery.dataTables.min.js"></script>
<script>
    $(document).ready(function() {{
        $('#excelTable').DataTable({{
            searching: true,
            ordering: true,
        }});
    }});
</script>
'''

                    # Write index.md
                    index_md_path = os.path.join(project_path, "index.md")
                    with open(index_md_path, "w", encoding="utf-8") as f:
                        f.write(md_content)

                    # Add project to vehicle nav
                    vehicle_nav.append({project: f"Vehicles/{vehicle}/{project}/index.md"})

            # Add vehicle to nav if it has projects
            if vehicle_nav:
                nav.append({vehicle: vehicle_nav})

    # Generate mkdocs.yml
    mkdocs_config = {
        "site_name": "Wire Harnessing Docs",
        "theme": {
            "name": "material",
            "palette": [
                {"scheme": "slate", "primary": "blue"}
            ],
            "favicon": "favicon.ico",
        },
        "nav": nav,
        "plugins": [
            "macros",
            "search",
        ],
        "markdown_extensions": ["attr_list"],
        "extra_javascript": [
            "https://code.jquery.com/jquery-3.6.0.min.js",
            "https://cdn.datatables.net/1.13.6/js/jquery.dataTables.min.js"
        ],
        "extra_css": ["https://cdn.datatables.net/1.13.6/css/jquery.dataTables.min.css"],
    }

    with open("mkdocs.yml", "w", encoding="utf-8") as f:
        yaml.dump(mkdocs_config, f, sort_keys=False)

    print("Auto-generated project pages with renamed files and mkdocs.yml successfully.")


if __name__ == "__main__":
    print("starting documentation generation...")
    delete_vehicle_subdirs()
    copy_projects()
    generate_docs_and_nav()
